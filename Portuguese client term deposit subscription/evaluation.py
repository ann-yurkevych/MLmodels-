import io
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import (
    f1_score, precision_score, recall_score,
    roc_auc_score, roc_curve, classification_report
)

def evaluate_model(estimator, X, y_true, model_name, method, preprocessed=False):
    """
    Returns a dict of metrics for one model.
    preprocessed=True  → X is already preprocessed (Hyperopt models)
    preprocessed=False → X goes through the full pipeline (RandomizedSearch models)
    """
    y_pred      = estimator.predict(X)
    y_pred_prob = estimator.predict_proba(X)[:, 1]
 
    fpr, tpr, _ = roc_curve(y_true, y_pred_prob)
 
    return {
        'model_name':  model_name,
        'method':      method,
        'precision':   precision_score(y_true, y_pred,      average='weighted'),
        'recall':      recall_score(y_true, y_pred,         average='weighted'),
        'f1':          f1_score(y_true, y_pred,             average='weighted'),
        'roc_auc':     roc_auc_score(y_true, y_pred_prob),
        'fpr':         fpr,
        'tpr':         tpr,
        'estimator':   estimator,
        'preprocessed': preprocessed,
    }

HEADER_FILL = PatternFill('solid', start_color='1F3864') 
ALT_FILL    = PatternFill('solid', start_color='D9E1F2')   # light blue
BEST_FILL   = PatternFill('solid', start_color='E2EFDA')   # light green
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT   = Font(name='Arial', size=10)
BOLD_FONT   = Font(name='Arial', bold=True, size=10)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left',   vertical='center')
_thin       = Side(style='thin', color='BFBFBF')
BORDER      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    cell.border = BORDER
 
 
def _style_body(cell, bold=False, fill=None):
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    if fill:
        cell.fill = fill

def plot_roc_curves(results):
    """
    Build a combined ROC curve figure for all results.
 
    Returns
    -------
    buf : BytesIO containing a PNG image (seeked to 0)
    """
    fig, ax = plt.subplots(figsize=(10, 7))
    colors_rs  = plt.cm.Blues(np.linspace(0.4, 0.9, 6))
    colors_hop = plt.cm.Oranges(np.linspace(0.4, 0.9, 3))
    ci_rs, ci_hop = 0, 0
 
    for r in results:
        if r['method'] == 'RandomizedSearchCV':
            color = colors_rs[ci_rs]; ci_rs += 1
        else:
            color = colors_hop[ci_hop]; ci_hop += 1
        ax.plot(
            r['fpr'], r['tpr'],
            label=f"{r['model_name']} ({r['method'][:5]}) AUC={r['roc_auc']:.3f}",
            color=color, linewidth=1.8,
        )
 
    ax.plot([0, 1], [0, 1], 'k--', linewidth=1, alpha=0.5, label='Random Classifier')
    ax.set_xlabel('False Positive Rate', fontsize=12)
    ax.set_ylabel('True Positive Rate',  fontsize=12)
    ax.set_title('ROC Curves — All Models', fontsize=14, fontweight='bold')
    ax.legend(loc='lower right', fontsize=8, framealpha=0.9)
    ax.grid(True, alpha=0.3)
    ax.set_xlim([0, 1]); ax.set_ylim([0, 1.02])
    plt.tight_layout()
 
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)
    return buf

 
def _build_summary_sheet(ws, results, best_result, best_models_randomized,
                          best_models_hyperopt, y_test, y_test_pred,
                          y_test_pred_prob):
    """Populate the Summary sheet."""
    ws.freeze_panes = 'A3'
 
    title_cell = ws['A1']
    title_cell.value     = 'Model Evaluation Results — Bank Marketing Dataset'
    title_cell.font      = Font(name='Arial', bold=True, size=14, color='1F3864')
    title_cell.alignment = LEFT
    ws.merge_cells('A1:I1')
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
 
    headers = ['Model', 'Tuning Method', 'Precision', 'Recall',
               'F1-Score', 'ROC AUC', 'Best Params', 'Val/Test', 'Notes']
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=2, column=col_idx, value=h))
 
    col_widths = [22, 20, 12, 12, 12, 12, 55, 10, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    best_key = (best_result['model_name'], best_result['method'])
 
    for row_idx, r in enumerate(results, start=3):
        is_best = (r['model_name'], r['method']) == best_key
        fill    = BEST_FILL if is_best else (ALT_FILL if row_idx % 2 == 0 else None)
 
        params_str = str(
            best_models_randomized.get(r['model_name'],
            best_models_hyperopt.get(r['model_name'], {})).get('params', '')
        )
        row_data = [
            r['model_name'], r['method'],
            round(r['precision'], 4), round(r['recall'], 4),
            round(r['f1'], 4),        round(r['roc_auc'], 4),
            params_str, 'Validation',
            '★ Best Model' if is_best else '',
        ]
        ws.row_dimensions[row_idx].height = 18
        for col_idx, val in enumerate(row_data, start=1):
            _style_body(ws.cell(row=row_idx, column=col_idx, value=val),
                        bold=is_best, fill=fill)
 
    # Test-set row for the best model
    test_row = len(results) + 3
    ws.row_dimensions[test_row].height = 18
    test_data = [
        best_result['model_name'], best_result['method'],
        round(precision_score(y_test, y_test_pred, average='weighted'), 4),
        round(recall_score(y_test,    y_test_pred, average='weighted'), 4),
        round(f1_score(y_test,        y_test_pred, average='weighted'), 4),
        round(roc_auc_score(y_test, y_test_pred_prob), 4),
        '', 'TEST SET', '★ Final Evaluation',
    ]
    for col_idx, val in enumerate(test_data, start=1):
        _style_body(
            ws.cell(row=test_row, column=col_idx, value=val),
            bold=True,
            fill=PatternFill('solid', start_color='FFE699'),
        )
 
 
def _build_roc_sheet(ws, results):
    """Populate the ROC Curves sheet."""
    ws['A1'].value = 'ROC Curves — All Models'
    ws['A1'].font  = Font(name='Arial', bold=True, size=13, color='1F3864')
 
    buf = plot_roc_curves(results)
    img = XLImage(buf)
    img.anchor = 'A3'
    ws.add_image(img)
 
 
def _build_reports_sheet(ws, results, X_val_proc, X_validation, y_validation):
    """Populate the Classification Reports sheet."""
    for c in ['A', 'B']:
        ws.column_dimensions[c].width = 22 if c == 'A' else 18
    for c in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 14
 
    ws['A1'].value = 'Per-Model Classification Reports (Validation Set)'
    ws['A1'].font  = Font(name='Arial', bold=True, size=13, color='1F3864')
 
    current_row = 3
    for r in results:
        # Model header
        cell = ws.cell(row=current_row, column=1,
                       value=f"{r['model_name']}  [{r['method']}]")
        cell.font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        cell.fill      = HEADER_FILL
        cell.alignment = LEFT
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row,   end_column=6)
        ws.row_dimensions[current_row].height = 20
        current_row += 1
 
        # Column sub-headers
        for ci, ch in enumerate(['Class', 'Precision', 'Recall',
                                  'F1-Score', 'Support', ''], start=1):
            _style_header(ws.cell(row=current_row, column=ci, value=ch),
                          fill=PatternFill('solid', start_color='2E75B6'))
        current_row += 1
 
        # Prediction & report
        y_pred_r = (r['estimator'].predict(X_val_proc)
                    if r['preprocessed']
                    else r['estimator'].predict(X_validation))
 
        report_dict = classification_report(y_validation, y_pred_r, output_dict=True)
        for class_label, metrics in report_dict.items():
            if class_label == 'accuracy' or not isinstance(metrics, dict):
                continue
            row_fill = ALT_FILL if current_row % 2 == 0 else None
            for ci, val in enumerate([
                class_label,
                round(metrics.get('precision', 0), 4),
                round(metrics.get('recall',    0), 4),
                round(metrics.get('f1-score',  0), 4),
                int(metrics.get('support',     0)),
                '',
            ], start=1):
                _style_body(ws.cell(row=current_row, column=ci, value=val),
                            fill=row_fill)
            current_row += 1
 
        # AUC row
        auc_cell = ws.cell(row=current_row, column=1,
                           value=f"ROC AUC: {r['roc_auc']:.4f}")
        auc_cell.font      = BOLD_FONT
        auc_cell.alignment = LEFT
        auc_cell.fill      = PatternFill('solid', start_color='FFF2CC')
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row,   end_column=6)
        current_row += 2   # blank line between models
 
def build_test_report_sheet(ws, best_result, X_test_proc, X_test, y_test):
    """Populate the Test Set Classification Report sheet for the best model only."""
    for c in ['A', 'B']:
        ws.column_dimensions[c].width = 22 if c == 'A' else 18
    for c in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 14

    ws['A1'].value = f'Test Set Report — Best Model: {best_result["model_name"]} ({best_result["method"]})'
    ws['A1'].font  = Font(name='Arial', bold=True, size=13, color='1F3864')

    current_row = 3

    # Model header
    cell = ws.cell(row=current_row, column=1,
                   value=f"{best_result['model_name']}  [{best_result['method']}]")
    cell.font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill      = HEADER_FILL
    cell.alignment = LEFT
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row,   end_column=6)
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # Column headers
    for ci, ch in enumerate(['Class', 'Precision', 'Recall',
                              'F1-Score', 'Support', ''], start=1):
        _style_header(ws.cell(row=current_row, column=ci, value=ch),
                      fill=PatternFill('solid', start_color='2E75B6'))
    current_row += 1

    # Predict on test set
    y_pred_test = (best_result['estimator'].predict(X_test_proc)
                   if best_result['preprocessed']
                   else best_result['estimator'].predict(X_test))

    report_dict = classification_report(y_test, y_pred_test, output_dict=True)
    for class_label, metrics in report_dict.items():
        if class_label == 'accuracy' or not isinstance(metrics, dict):
            continue
        row_fill = ALT_FILL if current_row % 2 == 0 else None
        for ci, val in enumerate([
            class_label,
            round(metrics.get('precision', 0), 4),
            round(metrics.get('recall',    0), 4),
            round(metrics.get('f1-score',  0), 4),
            int(metrics.get('support',     0)),
            '',
        ], start=1):
            _style_body(ws.cell(row=current_row, column=ci, value=val),
                        fill=row_fill)
        current_row += 1

    # AUC row
    y_pred_prob_test = (best_result['estimator'].predict_proba(X_test_proc)[:, 1]
                        if best_result['preprocessed']
                        else best_result['estimator'].predict_proba(X_test)[:, 1])
    auc_cell = ws.cell(row=current_row, column=1,
                       value=f"ROC AUC: {roc_auc_score(y_test, y_pred_prob_test):.4f}")
    auc_cell.font      = BOLD_FONT
    auc_cell.alignment = LEFT
    auc_cell.fill      = PatternFill('solid', start_color='FFF2CC')
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row,   end_column=6)
 
def build_excel_report(
    results,
    best_result,
    best_models_randomized,
    best_models_hyperopt,
    y_test,
    y_test_pred,
    y_test_pred_prob,
    X_val_proc,
    X_validation,
    y_validation,
    X_test_proc,
    X_test
    output_path='model_results.xlsx',
    ):
    """
    Build and save the full Excel workbook with three sheets:
      1. Summary          — all model metrics + test-set row for the winner
      2. ROC Curves       — combined ROC plot
      3. Classification Reports — per-model breakdown on the validation set
 
    Parameters
    ----------
    results               : list of dicts returned by evaluate_model()
    best_result           : single dict from results (the winning model)
    best_models_randomized: dict[name -> {estimator, params, method}]
    best_models_hyperopt  : dict[name -> {estimator, params, method}]
    y_test                : true test labels
    y_test_pred           : predicted test labels
    y_test_pred_prob      : predicted test probabilities (positive class)
    X_val_proc            : preprocessed validation features (for Hyperopt models)
    X_validation          : raw validation features (for pipeline models)
    y_validation          : true validation labels
    output_path           : file path to write the .xlsx
    """
    wb = openpyxl.Workbook()
 
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _build_summary_sheet(
        ws_summary, results, best_result,
        best_models_randomized, best_models_hyperopt,
        y_test, y_test_pred, y_test_pred_prob,
    )
 
    _build_roc_sheet(wb.create_sheet('ROC Curves'), results)
 
    _build_reports_sheet(
        wb.create_sheet('Classification Reports'),
        results, X_val_proc, X_validation, y_validation,
    )

    build_test_report_sheet(
        wb.create_sheet('Test Set Report'),
        best_result, X_test_proc, X_test, y_test,
    )

    wb.save(output_path)
    print(f"\n  Results saved {output_path}")


