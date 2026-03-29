import mlflow
import pandas as pd
import numpy as np
from imblearn.pipeline import Pipeline
from imblearn.over_sampling import SMOTENC
from imblearn.over_sampling import SMOTE
from sklearn.pipeline import Pipeline as SklearnPipeline
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import StandardScaler, OneHotEncoder, FunctionTransformer
from sklearn.impute import SimpleImputer
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import io
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


from preprocessing import (
    load_data,
    extract_target,
    split_train_test,
    encode_target, 
    numeric_categorical_features,
    drop_columns,
    unknown_values_replace,
    binarize_feature,
    get_categorical_indices
)

from classifiers import (
    base_decision_tree_model, 
    base_logistic_regression,
    base_KNeighborsClassifier,
    base_xgboost,
    base_lightgbm,
    base_catboost
)

from configurations import params_search
from hyperparameters_tuning import (
    randomized_search_cv,
    tune_xgb,
    tune_lightgbm,
    tune_catboost
)
from sklearn.metrics import (
    f1_score, precision_score, recall_score,
    roc_auc_score, roc_curve, classification_report
)
from evaluation import evaluate_model

"""PIPELINE
1. Load the data. 
2. Encode the target variable from string to numeric value. 
3. Drop features (based on prior EDA). 
4. Create new features if needed for the model.
5. Divide dataset into training, test, validation using stratified train_test_split. 
6. Extract target variable from each training, test, validation sets.

STEPS 7-14 WILL BE DONE IN THE PIPELINE, using ColumnTransformer, FunctionTransformer

7. Detect and impute missing values: no numeric missing values were detected, "unknown" categories will be replaced with SimpleImputer "most_frequent"
8. Encode categorical variables. 
9. Binarize previous feature: based on EDA I found out that subscription rate is high if previous > 1.
10. Scale numeric variables (fit + transform only on training, transform on validation + test sets)
11. Apply SMOTE techniques to handle class imbalance (only on the training data). 
12. For each model train the model with hyperparameters.
13. Pick the best hyperparameters (hyperparamters for each model with lowest validation error are the best for the model).
14. Train the best model using hyperparameters.

15. Evaluate on test set. 
16. Analysis of feature importance table.
17. Analysis of impact of features on predictions using SHAP library.

"""

print("Loading the data")

raw_df = load_data("data/bank-additional-full.csv")

# encode target variable from str to num
raw_df = encode_target(raw_df, 'y') 

# drop features: based on conducted EDA in exporation.ipynb 
raw_df = drop_columns(raw_df, ['duration', 'emp.var.rate', 'nr.employed']) 

# split dataset into train/test: 80/20
X_train, X_test = split_train_test(raw_df, 'y')

# split train dataset into train/validation: 60/20
X_train, X_validation = split_train_test(X_train, 'y')

# extract target variable from each set: training, validation, test: y_train, y_validation, y_test
X_train, y_train = extract_target(X_train, 'y') 
X_validation, y_validation = extract_target(X_validation, 'y') 
X_test, y_test = extract_target(X_test, 'y') 

# After extract_target, before pipeline
X_train, imputer = unknown_values_replace(X_train, ['default', 'education'])
X_validation = unknown_values_replace(X_validation, ['default', 'education'], imputer=imputer)
X_test = unknown_values_replace(X_test,       ['default', 'education'], imputer=imputer)

# determine numeric, categorical features
numeric_features, categorical_features = numeric_categorical_features(X_train)

# indices for category columns
columns_category_indices = get_categorical_indices(X_train, numeric_features)

# PIPELINE STARTS # 

models = {
    'DecisionTree':         base_decision_tree_model(),
    'KNN':                  base_KNeighborsClassifier(),
    'LogisticRegression':   base_logistic_regression(),
    'XGBoost':              base_xgboost(),
    'CatBoostClassifier':   base_catboost(),
    'LGBMClassifier':       base_lightgbm(),
}


def build_pipeline(model, numeric_features, categorical_features):
    """
    Builds a full imblearn Pipeline:
    preprocessor → SMOTENC → model
    
    resampling is used only for training set
    """

    numeric_transformer = SklearnPipeline(steps=[
        ('imputer', SimpleImputer(strategy='median')),
        ('scaler', StandardScaler())
    ])

 
    categorical_transformer = SklearnPipeline(steps=[
        ('imputer', SimpleImputer(strategy='most_frequent')),
        ('encoder', OneHotEncoder(handle_unknown='ignore', sparse_output=False))
    ])

   
    preprocessor = ColumnTransformer(transformers=[
        ('num', numeric_transformer, numeric_features),
        ('cat', categorical_transformer, categorical_features)
    ])


    return Pipeline(steps=[ # imbalanced pipeline, not scikit
        ('preprocessor', preprocessor),
        ('smote', SMOTE(random_state=42)),
        ('model', model)          
    ])



best_models_randomized = {}

for name, model in models.items():
    print(f"\nTuning of {name}")
    pipeline = build_pipeline(model, numeric_features, categorical_features)
    
    best_estimator, best_params = randomized_search_cv(
        pipeline,
        name, 
        X_train, y_train
    )
    best_models_randomized[name] = {
        'estimator': best_estimator,
        'params': best_params,
        'method': 'RandomizedSearchCV',
    }
    print(f"{name} best params: {best_params}")

# fitting preprocessing on training, transforming on test/validation
_prep_pipeline = build_pipeline(base_xgboost(), numeric_features, categorical_features)
_prep_pipeline.named_steps['preprocessor'].fit(X_train, y_train)
preprocessor_fit = _prep_pipeline.named_steps['preprocessor']
 
X_train_proc = preprocessor_fit.transform(X_train)
X_val_proc   = preprocessor_fit.transform(X_validation)
X_test_proc  = preprocessor_fit.transform(X_test)

# SMOTE on training only
smote = SMOTE(random_state=42)
X_train_resampled, y_train_resampled = smote.fit_resample(X_train_proc, y_train)
 
boosting_tuners = {
    'XGBoost':            tune_xgb,
    'LGBMClassifier':     tune_lightgbm,
    'CatBoostClassifier': tune_catboost,
}
 
best_models_hyperopt = {}

for name, tuner in boosting_tuners.items():
    print(f"\n  Tuning {name} with Hyperopt...")
    best_estimator, best_params = tuner(
        X_train_resampled, y_train_resampled,
        X_val_proc,        y_validation
    )
    best_models_hyperopt[name] = {
        'estimator': best_estimator,
        'params':    best_params,
        'method':    'Hyperopt',
    }
    print(f"  Best params: {best_params}")

results = []
for name, result in best_models_randomized.items():
    r = evaluate_model(
        result['estimator'], X_validation, y_validation,
        name, 'RandomizedSearchCV', preprocessed=False
    )
    results.append(r)
    print(f"  {name:<22} RandomizedSearchCV  F1={r['f1']:.4f}  AUC={r['roc_auc']:.4f}")
 
for name, result in best_models_hyperopt.items():
    r = evaluate_model(
        result['estimator'], X_val_proc, y_validation,
        name, 'Hyperopt', preprocessed=True
    )
    results.append(r)
    print(f"  {name:<22} Hyperopt  F1={r['f1']:.4f}  AUC={r['roc_auc']:.4f}")


best_result = max(results, key=lambda r: r['roc_auc'])
print(f"\n  ★ Best model: {best_result['model_name']} ({best_result['method']})")
print(f"    Validation AUC: {best_result['roc_auc']:.4f}")
 
if best_result['preprocessed']:
    y_test_pred      = best_result['estimator'].predict(X_test_proc)
    y_test_pred_prob = best_result['estimator'].predict_proba(X_test_proc)[:, 1]
else:
    y_test_pred      = best_result['estimator'].predict(X_test)
    y_test_pred_prob = best_result['estimator'].predict_proba(X_test)[:, 1]
 
test_f1  = f1_score(y_test, y_test_pred, average='weighted')
test_auc = roc_auc_score(y_test, y_test_pred_prob)
 
print(f"\n  FINAL TEST RESULTS")
print(f"  F1 (weighted): {test_f1:.4f}")
print(f"  ROC AUC:       {test_auc:.4f}")
print(f"\n{classification_report(y_test, y_test_pred)}")


wb = openpyxl.Workbook()
 
# ── colour palette ──────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', start_color='1F3864')   # dark navy
ALT_FILL      = PatternFill('solid', start_color='D9E1F2')   # light blue
BEST_FILL     = PatternFill('solid', start_color='E2EFDA')   # light green
HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT     = Font(name='Arial', size=10)
BOLD_FONT     = Font(name='Arial', bold=True, size=10)
CENTER        = Alignment(horizontal='center', vertical='center')
LEFT          = Alignment(horizontal='left',   vertical='center')
thin          = Side(style='thin', color='BFBFBF')
BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)
 
 
def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill      = fill
    cell.font      = font
    cell.alignment = CENTER
    cell.border    = BORDER
 
 
def style_body(cell, bold=False, fill=None):
    cell.font      = BOLD_FONT if bold else BODY_FONT
    cell.alignment = CENTER
    cell.border    = BORDER
    if fill:
        cell.fill = fill
 
 

ws = wb.active
ws.title = 'Summary'
ws.freeze_panes = 'A3'
 
title_cell = ws['A1']
title_cell.value     = 'Model Evaluation Results — Bank Marketing Dataset'
title_cell.font      = Font(name='Arial', bold=True, size=14, color='1F3864')
title_cell.alignment = LEFT
ws.merge_cells('A1:I1')
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 22
 
headers = ['Model', 'Tuning Method', 'Precision', 'Recall',
           'F1-Score', 'ROC AUC', 'Best Params', 'Val/Test', 'Notes']
 
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    style_header(cell)
 
col_widths = [22, 20, 12, 12, 12, 12, 55, 10, 20]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
 
best_key = (best_result['model_name'], best_result['method'])
 
for row_idx, r in enumerate(results, start=3):
    is_best = (r['model_name'], r['method']) == best_key
    fill    = BEST_FILL if is_best else (ALT_FILL if row_idx % 2 == 0 else None)
 
    row_data = [
        r['model_name'],
        r['method'],
        round(r['precision'], 4),
        round(r['recall'],    4),
        round(r['f1'],        4),
        round(r['roc_auc'],   4),
        str(best_models_randomized.get(r['model_name'],
            best_models_hyperopt.get(r['model_name'], {})).get('params', '')),
        'Validation',
        '★ Best Model' if is_best else '',
    ]
    ws.row_dimensions[row_idx].height = 18
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        style_body(cell, bold=is_best, fill=fill)
 
# Test row for best model
test_row = len(results) + 3
ws.row_dimensions[test_row].height = 18
test_data = [
    best_result['model_name'],
    best_result['method'],
    round(precision_score(y_test, y_test_pred, average='weighted'), 4),
    round(recall_score(y_test, y_test_pred,    average='weighted'), 4),
    round(test_f1,  4),
    round(test_auc, 4),
    '',
    'TEST SET',
    '★ Final Evaluation',
]
for col_idx, val in enumerate(test_data, start=1):
    cell = ws.cell(row=test_row, column=col_idx, value=val)
    style_body(cell, bold=True, fill=PatternFill('solid', start_color='FFE699'))
 
 
# AUC ROC
ws_roc = wb.create_sheet('ROC Curves')
ws_roc['A1'].value = 'ROC Curves — All 12 Models'
ws_roc['A1'].font  = Font(name='Arial', bold=True, size=13, color='1F3864')
 
# One combined ROC plot
fig, ax = plt.subplots(figsize=(10, 7))
colors_rs  = plt.cm.Blues(np.linspace(0.4, 0.9, 6))
colors_hop = plt.cm.Oranges(np.linspace(0.4, 0.9, 3))
ci_rs, ci_hop = 0, 0
 
for r in results:
    if r['method'] == 'RandomizedSearchCV':
        color = colors_rs[ci_rs]; ci_rs += 1
    else:
        color = colors_hop[ci_hop]; ci_hop += 1
    ax.plot(r['fpr'], r['tpr'],
            label=f"{r['model_name']} ({r['method'][:5]}) AUC={r['roc_auc']:.3f}",
            color=color, linewidth=1.8)
 
ax.plot([0, 1], [0, 1], 'k--', linewidth=1, alpha=0.5, label='Random Classifier')
ax.set_xlabel('False Positive Rate', fontsize=12)
ax.set_ylabel('True Positive Rate', fontsize=12)
ax.set_title('ROC Curves — All 12 Models', fontsize=14, fontweight='bold')
ax.legend(loc='lower right', fontsize=8, framealpha=0.9)
ax.grid(True, alpha=0.3)
ax.set_xlim([0, 1]); ax.set_ylim([0, 1.02])
plt.tight_layout()
 
buf = io.BytesIO()
fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
buf.seek(0)
img = XLImage(buf)
img.anchor = 'A3'
ws_roc.add_image(img)
plt.close(fig)
 
 
# ── Sheet 3: Per-model Classification Reports ───────────────
ws_report = wb.create_sheet('Classification Reports')
ws_report.column_dimensions['A'].width = 22
ws_report.column_dimensions['B'].width = 18
for c in ['C', 'D', 'E', 'F']:
    ws_report.column_dimensions[c].width = 14
 
ws_report['A1'].value = 'Per-Model Classification Reports (Validation Set)'
ws_report['A1'].font  = Font(name='Arial', bold=True, size=13, color='1F3864')
 
current_row = 3
for r in results:
    # Model header
    header_cell = ws_report.cell(row=current_row, column=1,
                                 value=f"{r['model_name']}  [{r['method']}]")
    header_cell.font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_cell.fill      = HEADER_FILL
    header_cell.alignment = LEFT
    ws_report.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row,   end_column=6
    )
    ws_report.row_dimensions[current_row].height = 20
    current_row += 1
 
    # Column headers
    col_headers = ['Class', 'Precision', 'Recall', 'F1-Score', 'Support', '']
    for ci, ch in enumerate(col_headers, start=1):
        cell = ws_report.cell(row=current_row, column=ci, value=ch)
        style_header(cell, fill=PatternFill('solid', start_color='2E75B6'))
    current_row += 1
 
    # Metric rows
    if r['preprocessed']:
        y_pred_r = r['estimator'].predict(X_val_proc)
    else:
        y_pred_r = r['estimator'].predict(X_validation)
 
    report_dict = classification_report(y_validation, y_pred_r, output_dict=True)
    for class_label, metrics in report_dict.items():
        if class_label in ('accuracy',):
            continue
        if isinstance(metrics, dict):
            row_fill = ALT_FILL if current_row % 2 == 0 else None
            for ci, val in enumerate([
                class_label,
                round(metrics.get('precision', 0), 4),
                round(metrics.get('recall',    0), 4),
                round(metrics.get('f1-score',  0), 4),
                int(metrics.get('support',     0)),
                '',
            ], start=1):
                cell = ws_report.cell(row=current_row, column=ci, value=val)
                style_body(cell, fill=row_fill)
            current_row += 1
 
    # AUC row
    auc_cell = ws_report.cell(row=current_row, column=1,
                               value=f"ROC AUC: {r['roc_auc']:.4f}")
    auc_cell.font      = BOLD_FONT
    auc_cell.alignment = LEFT
    auc_cell.fill      = PatternFill('solid', start_color='FFF2CC')
    ws_report.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row,   end_column=6
    )
    current_row += 2   # blank row between models
 
 
output_path = '/mnt/user-data/outputs/model_results.xlsx'
wb.save(output_path)
print(f"\n  Results saved → {output_path}")